from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
CORS(app)

FILE_NAME = "students.xlsx"

# Create file if not exists
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.append(["First Name", "Last Name", "Grade", "DOB", "Parent Name", "Phone", "Email"])
    wb.save(FILE_NAME)

@app.route('/submit', methods=['POST'])
def submit():
    data = request.json

    # Load excel
    wb = load_workbook(FILE_NAME)
    ws = wb.active

    # Add row
    ws.append([
        data.get("firstName"),
        data.get("lastName"),
        data.get("grade"),
        data.get("dob"),
        data.get("parentName"),
        data.get("phone"),
        data.get("email")
    ])

    wb.save(FILE_NAME)

    return jsonify({"message": "Saved to Excel ✅"})

if __name__ == '__main__':
    app.run(debug=True)